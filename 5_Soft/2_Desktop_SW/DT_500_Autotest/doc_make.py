# import docx
from openpyxl import Workbook, load_workbook
import time

class Doc_Report(object):
    '''отчет и Word, Excel '''

    def __init__(self):
        # шаблон
        self.template_path = "Template/Template_report.xlsx"

    def make_report_word(self, list):
        '''создание отчета'''
        pass

    def make_report_excel(self, measure_data):
        '''создание отчета
        структура list
        [№ датчика, статус, Uпит, Iпит,[Uизм(0А), Uизм(100А), Uизм(200А), Uизм(300А), Uизм(400А), Uизм(500А)]]]
        '''
        report = load_workbook(self.template_path)

        wsheet_1 = report[report.sheetnames[0]]
        wsheet_2 = report[report.sheetnames[1]]

        # заполнение 1 листа
        for num_1, item in enumerate(measure_data):
            for num_2, data in enumerate(item):
                if num_2 < 4:
                    wsheet_1.cell(row=num_1+2, column=num_2+2).value = str(data)
            for num_3, data in enumerate(item[4]):
                wsheet_1.cell(row=num_1 + 2, column= 6 + num_3).value = str(data)


        report.save(time.strftime("Протоколы\Протокол от %H_%M %d.%m.%Y.xlsx"))
        # report.close()

        pass

    def set_header(self, excel_sheet, header_list):
        for num, name_header in enumerate(header_list):
            excel_sheet.cell(row=1, column=num + 1, value=name_header)

    def debug_make_template_report_excel(self):
        wb = Workbook()

        wsheet_2 = wb.create_sheet("Данные", 1)
        wsheet_3 = wb.create_sheet("Доп. данные", 2)

        # заполняем первую страницу шаблона - это итоги работы общие
        # заголовок
        header_list_1 = ["Номер датчика", "Исправность датчика", "Условия отказа"]
        self.set_header(wsheet_1, header_list_1)

        for x in range(1, 33):
            wsheet_1.cell(row=x+1, column=1, value=x)


        # заполняем первую страницу шаблона - это детально
        # заголовок
        header_list_2 = ["Группа датчиков",
                         "Номер датчика",
                         "Исправность датчика",
                         "Uпит, В",
                         "Iпит, mA",
                         "Uизм, В (0 А)",
                         "Uизм, В (100 А)",
                         "Uизм, В (200 А)",
                         "Uизм, В (300 А)",
                         "Uизм, В (400 А)",
                         "Uизм, В (500 А)"]

        self.set_header(wsheet_2, header_list_2)

        k = 1
        for i in range(1, 33*8, 8*3):
            wsheet_2.cell(row=i+1, column=1, value=k)
            k += 1

        num_sens = 1
        for x in range(1, (33*3)-1, 3):
            wsheet_2.cell(row=x+1, column=2, value=num_sens)
            wsheet_2.cell(row=x+1, column=4, value=16.0)
            wsheet_2.cell(row=x+2, column=4, value=27.0)
            wsheet_2.cell(row=x+3, column=4, value=32.0)
            num_sens += 1

        # третья страница - это лог работы
        # заголовок
        self.set_header(wsheet_3, ["Время", "Действие"])

        wb.save("Debug.xlsx")
