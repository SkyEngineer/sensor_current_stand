# import docx
from openpyxl import Workbook, load_workbook
import time
import os

class Doc_Report(object):
    '''отчет и Word, Excel '''

    def __init__(self):
        # шаблон
        self.template_path = "Template/Template_report.xlsx"
        self.doc_name = ""

    def make_report_word(self, list):
        '''создание отчета'''
        pass

    def make_report_excel(self, measure_data, name):
        '''создание отчета
        структура list
        [№ датчика, статус, Uпит, Iпит,[Uизм(0А), Uизм(100А), Uизм(200А), Uизм(300А), Uизм(400А), Uизм(500А)]]]
        '''

        # name = time.strftime("Протоколы\Протокол от %d.%m.%Y в %H_%M_%S.xlsx")
        if name.split("\\")[-1] in os.listdir(name.split("\\")[0]):
            report = load_workbook(name)
        else:
            report = load_workbook(self.template_path)

        wsheet_1 = report[report.sheetnames[0]]
        empty_row = 0
        # заполнение 1 листа
        # определение уже заполненных строк
        for i in range(1, 73):
            if wsheet_1.cell(row=i, column=1).value == None:
                empty_row = i
                break


        for num_1, item_1 in enumerate(measure_data):
            item = item_1[0:4] + item_1[4]
            #print(f"item_1 {item}")
            #item = item_1[0:4] + item_1[-1]
            #print(item)

            row_temp = empty_row + num_1
            for k, data in enumerate(item):
                wsheet_1.cell(row=row_temp, column=k+1).value = str(data)
                #print(f"data {data}")

        report.save(name)
        # report.close()

        pass

    def set_header(self, excel_sheet, header_list):
        for num, name_header in enumerate(header_list):
            excel_sheet.cell(row=1, column=num + 1, value=name_header)

    def debug_make_template_report_excel(self):
        wb = Workbook()

        wsheet_2 = wb.create_sheet("Данные", 1)
        wsheet_3 = wb.create_sheet("Доп. данные", 2)

        # заполняем первую страницу шаблона - это итоги работы общие
        # заголовок
        header_list_1 = ["Номер датчика", "Исправность датчика", "Условия отказа"]
        self.set_header(wsheet_1, header_list_1)

        for x in range(1, 33):
            wsheet_1.cell(row=x+1, column=1, value=x)


        # заполняем первую страницу шаблона - это детально
        # заголовок
        header_list_2 = ["Группа датчиков",
                         "Номер датчика",
                         "Исправность датчика",
                         "Uпит, В",
                         "Iпит, mA",
                         "Uизм, В (0 А)",
                         "Uизм, В (100 А)",
                         "Uизм, В (200 А)",
                         "Uизм, В (300 А)",
                         "Uизм, В (400 А)",
                         "Uизм, В (500 А)"]

        self.set_header(wsheet_2, header_list_2)

        k = 1
        for i in range(1, 33*8, 8*3):
            wsheet_2.cell(row=i+1, column=1, value=k)
            k += 1

        num_sens = 1
        for x in range(1, (33*3)-1, 3):
            wsheet_2.cell(row=x+1, column=2, value=num_sens)
            wsheet_2.cell(row=x+1, column=4, value=16.0)
            wsheet_2.cell(row=x+2, column=4, value=27.0)
            wsheet_2.cell(row=x+3, column=4, value=32.0)
            num_sens += 1

        # третья страница - это лог работы
        # заголовок
        self.set_header(wsheet_3, ["Время", "Действие"])

        wb.save("Debug.xlsx")
